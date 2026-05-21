import numpy as np
import pandas as pd
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE  = "../output/top20_accounts_seasonality.xlsx"
OUTPUT_FILE = "../output/top20_accounts_velocity.xlsx"

# Velocity thresholds
HIGH_BASELINE = 50
HIGH_SELL_PCT = 0.60
MOD_BASELINE  = 15
MOD_SELL_PCT  = 0.40
DORMANT_PCT   = 0.10

COLORS = {
    "High Velocity":     "00B050",
    "Moderate Velocity": "FFC000",
    "Low Velocity":      "FF6600",
    "Dormant":           "FF0000",
}


def classify(baseline, sell_pct):
    if sell_pct < DORMANT_PCT:
        return "Dormant"
    if baseline >= HIGH_BASELINE and sell_pct >= HIGH_SELL_PCT:
        return "High Velocity"
    if baseline >= MOD_BASELINE and sell_pct >= MOD_SELL_PCT:
        return "Moderate Velocity"
    return "Low Velocity"


def compute_baseline(values):
    selling = values[values > 0]
    if len(selling) == 0:
        return 0.0
    cap   = np.quantile(selling, 0.90)
    floor = np.quantile(selling, 0.10)
    return round(float(np.clip(selling, floor, cap).mean()), 2)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Color all VC_ columns and VELOCITY_OVERALL
    col_indices = {
        i: cell.value for i, cell in enumerate(ws[1], 1)
        if cell.value and (str(cell.value).startswith("VC_") or cell.value == "VELOCITY_OVERALL")
    }
    for row in ws.iter_rows(min_row=2):
        for col_idx in col_indices:
            cell  = row[col_idx - 1]
            color = COLORS.get(cell.value)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def main():
    xl        = pd.ExcelFile(INPUT_FILE)
    meta_cols = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

    # ── Load all sheets, track account → item mapping ─────────────────────────
    frames           = []
    account_item_map = defaultdict(set)

    for sheet in xl.sheet_names:
        df        = xl.parse(sheet)
        week_cols_sheet = [
            c for c in df.columns
            if isinstance(c, str) and c[0].isdigit() and "W" in c and not c.startswith("SB")
        ]
        keep = [c for c in meta_cols if c in df.columns] + week_cols_sheet
        frames.append(df[keep])
        for item_id in df["ITEM_ID"].unique():
            account_item_map[sheet].add(item_id)

    all_data  = pd.concat(frames, ignore_index=True)
    week_cols = [
        c for c in all_data.columns
        if isinstance(c, str) and c[0].isdigit() and "W" in c
    ]
    all_years = sorted(set(w.split("-")[0] for w in week_cols))

    # ── Aggregate sales per item across all accounts ──────────────────────────
    meta = all_data[meta_cols].drop_duplicates(subset=["ITEM_ID"])
    agg  = all_data.groupby("ITEM_ID")[week_cols].sum().reset_index()
    agg  = meta.merge(agg, on="ITEM_ID", how="right")

    # ── Classify velocity per item per year + overall ─────────────────────────
    vc_year_cols = [f"VC_{yr}" for yr in all_years]
    item_vc_rows = []

    for _, item_row in agg.iterrows():
        row = {c: item_row[c] for c in meta_cols}

        # Overall — all weeks combined
        all_qty      = item_row[week_cols].to_numpy(dtype=float)
        overall_base = compute_baseline(all_qty)
        overall_pct  = round((all_qty > 0).sum() / len(week_cols), 4)
        row["VELOCITY_OVERALL"] = classify(overall_base, overall_pct)

        # Per year
        for yr in all_years:
            yr_cols = [w for w in week_cols if w.startswith(yr + "-")]
            if not yr_cols:
                row[f"VC_{yr}"] = "No Data"
                continue
            yr_qty   = item_row[yr_cols].to_numpy(dtype=float)
            yr_base  = compute_baseline(yr_qty)
            yr_pct   = round((yr_qty > 0).sum() / len(yr_cols), 4)
            row[f"VC_{yr}"] = classify(yr_base, yr_pct)

        item_vc_rows.append(row)

    item_vc = pd.DataFrame(item_vc_rows)

    # ── Build item × account output ───────────────────────────────────────────
    account_rows = [
        {"ACCOUNT": acct, "ITEM_ID": item_id}
        for acct, item_ids in account_item_map.items()
        for item_id in item_ids
    ]
    account_df = pd.DataFrame(account_rows)
    result = account_df.merge(item_vc, on="ITEM_ID", how="left")
    result = result[["ACCOUNT"] + meta_cols + vc_year_cols + ["VELOCITY_OVERALL"]]
    result = result.sort_values(["ACCOUNT", "ITEM_ID"]).reset_index(drop=True)

    # ── Excel output ──────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="VELOCITY", index=False)
        style_sheet(writer.sheets["VELOCITY"])

        # Counts per year per class (item-level, deduplicated)
        count_rows = []
        for yr in all_years:
            col = f"VC_{yr}"
            for status, cnt in item_vc[col].value_counts().items():
                count_rows.append({"YEAR": int(yr), "VELOCITY_CLASS": status, "COUNT": cnt})
        counts = pd.DataFrame(count_rows).pivot_table(
            index="YEAR", columns="VELOCITY_CLASS", values="COUNT", fill_value=0
        ).reset_index()
        for s in ["High Velocity", "Moderate Velocity", "Low Velocity", "Dormant"]:
            if s not in counts.columns:
                counts[s] = 0
        counts = counts[["YEAR"] + [s for s in ["High Velocity", "Moderate Velocity", "Low Velocity", "Dormant"] if s in counts.columns]]
        counts.to_excel(writer, sheet_name="VELOCITY_COUNTS", index=False)
        style_sheet(writer.sheets["VELOCITY_COUNTS"])

    print(f"Output saved: {OUTPUT_FILE}")
    print(f"Rows (item × account): {len(result)}")
    print(f"Unique items: {item_vc['ITEM_ID'].nunique()}")
    print("\nOverall velocity breakdown:")
    print(item_vc["VELOCITY_OVERALL"].value_counts().to_string())
    print("\nBy year:")
    for yr in all_years:
        col = f"VC_{yr}"
        vc  = item_vc[col].value_counts()
        print(f"  {yr}: {dict(vc)}")


if __name__ == "__main__":
    main()
