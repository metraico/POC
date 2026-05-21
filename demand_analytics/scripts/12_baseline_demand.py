"""
12_baseline_demand.py

Generates a standardized weekly baseline demand dataset from
top20_accounts_consolidated.xlsx.

Baseline method — trimmed mean (retail forecasting standard):
  1. Aggregate daily → weekly actual demand
  2. For each item × account × year, find the 90th percentile of
     non-zero weeks → this is the spike cap (upper limit)
  3. Cap each week's actual at that limit (promo spikes are flattened)
  4. BASELINE_DEMAND = mean of capped values across all 52 weeks

This removes promotional spikes while preserving normal demand levels.
Simple yearly_total / 52 is NOT used — it distorts baselines when spikes exist.

Outputs (to account_consolidator/output/):
  - baseline_demand.parquet   ← for Streamlit
  - baseline_demand.xlsx      ← styled Excel, single sheet "Baseline Demand"

Run from Dataprocessing/:
    python3 12_baseline_demand.py
"""

import os
import time
import itertools
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE, "output", "top20_accounts_consolidated.xlsx")
OUTPUT_DIR = os.path.join(BASE, "output")
OUT_PARQ   = os.path.join(OUTPUT_DIR, "baseline_demand.parquet")
OUT_XLSX   = os.path.join(OUTPUT_DIR, "baseline_demand.xlsx")

INFO_COLS = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]
YEARS     = [2023, 2024, 2025, 2026]
WEEKS     = list(range(1, 53))  # 1–52

SPIKE_PERCENTILE = 0.90   # weeks above this quantile are considered promo spikes

OUTPUT_COLS = [
    "ACCOUNT", "ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM",
    "YEAR", "WEEK_NUM", "ACTUAL_WEEKLY_DEMAND", "SPIKE_CAP",
    "CAPPED_DEMAND", "BASELINE_DEMAND",
]

# ── Styling (matches consolidate_by_account.py conventions) ───────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CENTER      = Alignment(horizontal="center")
FILL_ODD    = PatternFill("solid", fgColor="DCE6F1")
FILL_EVEN   = PatternFill("solid", fgColor="FFFFFF")


def log(msg: str, level: str = "INFO"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [{level}] {msg}")


def section(title: str):
    log("─" * 60)
    log(f"  STEP: {title}")
    log("─" * 60)


# ── Step 1: Load and melt all sheets ─────────────────────────────────────────
section("Load & Melt All Sheets (wide → long)")
t0 = time.time()

xl = pd.ExcelFile(SOURCE)
log(f"Source: {SOURCE}")
log(f"Sheets: {xl.sheet_names}")

all_long = []
for sheet in xl.sheet_names:
    df_wide   = xl.parse(sheet)
    date_cols = [c for c in df_wide.columns if c not in INFO_COLS and c != "GRAND_TOTAL"]
    avail_info = [c for c in INFO_COLS if c in df_wide.columns]

    df_long = df_wide[avail_info + date_cols].melt(
        id_vars=avail_info, var_name="DATE", value_name="QTY"
    )

    df_long["ACCOUNT"] = sheet.title()
    df_long["DATE"]    = pd.to_datetime(df_long["DATE"], errors="coerce")
    df_long["ITEM_ID"] = pd.to_numeric(df_long["ITEM_ID"], errors="coerce").fillna(0).astype(int)
    df_long["QTY"]     = pd.to_numeric(df_long["QTY"], errors="coerce").fillna(0).astype(int)

    for col in avail_info:
        if col != "ITEM_ID":
            df_long[col] = df_long[col].astype(str).str.strip().replace("nan", "")

    df_long = df_long.dropna(subset=["DATE"])

    iso = df_long["DATE"].dt.isocalendar()
    df_long["ISO_YEAR"] = iso["year"].astype(int)
    df_long["ISO_WEEK"] = iso["week"].astype(int)

    all_long.append(df_long)
    log(f"  {sheet:<30} {len(df_wide):>5} items × {len(date_cols):>4} dates")

combined = pd.concat(all_long, ignore_index=True)
log(f"Combined rows: {len(combined):,}  ({time.time() - t0:.1f}s)")


# ── Step 2: Filter to target years, cap at week 52 ────────────────────────────
section("Filter Years & Cap at Week 52")
t0 = time.time()
combined = combined[
    combined["ISO_YEAR"].isin(YEARS) &
    (combined["ISO_WEEK"] <= 52)
].copy()
log(f"Rows after filter: {len(combined):,}  ({time.time() - t0:.1f}s)")


# ── Step 3: Aggregate daily → weekly ─────────────────────────────────────────
section("Aggregate Daily → Weekly")
t0 = time.time()
weekly_actuals = (
    combined
    .groupby(
        ["ACCOUNT", "ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM",
         "ISO_YEAR", "ISO_WEEK"],
        as_index=False, observed=True,
    )["QTY"]
    .sum()
    .rename(columns={"ISO_YEAR": "YEAR", "ISO_WEEK": "WEEK_NUM", "QTY": "ACTUAL_WEEKLY_DEMAND"})
)
log(f"Weekly actuals rows: {len(weekly_actuals):,}  ({time.time() - t0:.1f}s)")


# ── Step 4: Build scaffold (every item × account × year × week) ───────────────
section("Build Weekly Scaffold")
t0 = time.time()
item_universe = (
    combined[["ACCOUNT", "ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]]
    .drop_duplicates()
    .reset_index(drop=True)
)
year_week_grid = pd.DataFrame(
    list(itertools.product(YEARS, WEEKS)),
    columns=["YEAR", "WEEK_NUM"],
)
scaffold = item_universe.merge(year_week_grid, how="cross")
log(f"Item universe: {len(item_universe):,} unique item-account combinations")
log(f"Scaffold rows: {len(scaffold):,}  ({time.time() - t0:.1f}s)")


# ── Step 5: Left-join actuals onto scaffold ───────────────────────────────────
section("Join Actuals onto Scaffold")
t0 = time.time()
baseline = scaffold.merge(
    weekly_actuals[["ACCOUNT", "ITEM_ID", "YEAR", "WEEK_NUM", "ACTUAL_WEEKLY_DEMAND"]],
    on=["ACCOUNT", "ITEM_ID", "YEAR", "WEEK_NUM"],
    how="left",
)
baseline["ACTUAL_WEEKLY_DEMAND"] = baseline["ACTUAL_WEEKLY_DEMAND"].fillna(0).astype(int)
log(f"Baseline rows: {len(baseline):,}  ({time.time() - t0:.1f}s)")


# ── Step 6: Compute Baseline using Trimmed Mean (90th percentile cap) ─────────
#
# Logic:
#   A. For each (ACCOUNT, ITEM_ID, YEAR): compute the 90th percentile of
#      non-zero weekly actuals → SPIKE_CAP (the upper limit)
#      - Using only non-zero weeks avoids zero-inflation pulling the cap down
#        for items that only sell a few weeks per year
#   B. CAPPED_DEMAND = min(ACTUAL_WEEKLY_DEMAND, SPIKE_CAP) per week
#      - Promo spikes are flattened to the cap; normal weeks are unchanged
#   C. BASELINE_DEMAND = mean of CAPPED_DEMAND across all 52 weeks
#      - Zeros (non-selling weeks) ARE included so the baseline reflects
#        realistic average weekly demand including off-weeks
#
section("Compute Baseline Demand (trimmed mean — 90th percentile cap)")
t0 = time.time()

# A — 90th percentile of non-zero weeks per item × account × year
nonzero = baseline[baseline["ACTUAL_WEEKLY_DEMAND"] > 0]
spike_caps = (
    nonzero
    .groupby(["ACCOUNT", "ITEM_ID", "YEAR"], as_index=False)["ACTUAL_WEEKLY_DEMAND"]
    .quantile(SPIKE_PERCENTILE)
    .rename(columns={"ACTUAL_WEEKLY_DEMAND": "SPIKE_CAP"})
)
spike_caps["SPIKE_CAP"] = spike_caps["SPIKE_CAP"].round(2)

baseline = baseline.merge(
    spike_caps, on=["ACCOUNT", "ITEM_ID", "YEAR"], how="left"
)
# Items with no non-zero weeks at all get a cap of 0
baseline["SPIKE_CAP"] = baseline["SPIKE_CAP"].fillna(0).round(2)

# B — Cap each week's actual at the spike cap
baseline["CAPPED_DEMAND"] = baseline[["ACTUAL_WEEKLY_DEMAND", "SPIKE_CAP"]].min(axis=1).astype(int)

# C — Baseline = mean of capped demand across all 52 weeks (zeros included)
baseline_avg = (
    baseline
    .groupby(["ACCOUNT", "ITEM_ID", "YEAR"], as_index=False)["CAPPED_DEMAND"]
    .mean()
    .rename(columns={"CAPPED_DEMAND": "BASELINE_DEMAND"})
)
baseline_avg["BASELINE_DEMAND"] = baseline_avg["BASELINE_DEMAND"].round(2)

baseline = baseline.merge(
    baseline_avg, on=["ACCOUNT", "ITEM_ID", "YEAR"], how="left"
)

spikes_found = (baseline["ACTUAL_WEEKLY_DEMAND"] > baseline["SPIKE_CAP"]).sum()
log(f"Spike cap percentile : {int(SPIKE_PERCENTILE*100)}th")
log(f"Weeks capped (spikes): {spikes_found:,}")
log(f"Baseline computed    ({time.time() - t0:.1f}s)")


# ── Step 7: Final column order and sort ───────────────────────────────────────
baseline = (
    baseline[OUTPUT_COLS]
    .sort_values(["ACCOUNT", "ITEM_ID", "YEAR", "WEEK_NUM"])
    .reset_index(drop=True)
)
log(f"Final dataset: {len(baseline):,} rows × {len(OUTPUT_COLS)} columns")
log(f"Non-zero demand rows: {(baseline['ACTUAL_WEEKLY_DEMAND'] > 0).sum():,}")


# ── Step 8: Write Parquet ─────────────────────────────────────────────────────
section("Write Parquet")
t0 = time.time()
baseline.to_parquet(OUT_PARQ, index=False, engine="pyarrow")
log(f"Saved: {OUT_PARQ}  ({os.path.getsize(OUT_PARQ)/1024:.1f} KB)  ({time.time() - t0:.1f}s)")


# ── Step 9: Write Styled Excel — one sheet per account, wide format ───────────
#
# Structure mirrors top20_accounts_consolidated.xlsx:
#   Rows    = items (ITEM_ID, ITEM_DESCR, BRAND, VENDOR, PRODFAM)
#   Columns = W2023-01 … W2026-52  (one column per YEAR-WEEK)
#             + BASELINE_DEMAND column (the trimmed mean for that item × year)
#   One sheet per account
#
section("Write Excel — wide format, one sheet per account")
t0 = time.time()

WEEK_FILL   = PatternFill("solid", fgColor="BDD7EE")   # light blue for baseline col
ZERO_FILL   = PatternFill("solid", fgColor="F2F2F2")   # grey for zero cells
TOTAL_FONT  = Font(bold=True, size=10)

INFO_OUT = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

wb = openpyxl.Workbook()
wb.remove(wb.active)

accounts = baseline["ACCOUNT"].unique()
for acct in sorted(accounts):
    df_acct = baseline[baseline["ACCOUNT"] == acct].copy()

    # Build WEEK_LABEL for column names: "2023-W01", "2023-W02", …
    df_acct["WEEK_LABEL"] = (
        df_acct["YEAR"].astype(str) + "-W" +
        df_acct["WEEK_NUM"].astype(str).str.zfill(2)
    )

    # Pivot: rows = items, columns = week labels, values = ACTUAL_WEEKLY_DEMAND
    pivot = df_acct.pivot_table(
        index=INFO_OUT,
        columns="WEEK_LABEL",
        values="ACTUAL_WEEKLY_DEMAND",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    week_cols = [c for c in pivot.columns if c not in INFO_OUT]

    # Add per-item BASELINE_DEMAND (same value for all weeks of a year —
    # take the mean across weeks as one representative value per item)
    baseline_per_item = (
        df_acct.groupby(INFO_OUT)["BASELINE_DEMAND"]
        .mean().round(2).reset_index()
    )
    pivot = pivot.merge(baseline_per_item, on=INFO_OUT, how="left")

    all_cols   = INFO_OUT + week_cols + ["BASELINE_DEMAND"]
    baseline_col_idx = len(INFO_OUT) + len(week_cols) + 1  # 1-based

    ws = wb.create_sheet(title=acct[:31])

    # Header
    for ci, h in enumerate(all_cols, 1):
        cell           = ws.cell(row=1, column=ci, value=str(h))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # Data rows
    for ri, row in enumerate(pivot[all_cols].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == baseline_col_idx:
                cell.fill = WEEK_FILL
                cell.font = TOTAL_FONT
            elif ci > len(INFO_OUT) and (val == 0 or val == "0"):
                cell.fill = ZERO_FILL

    ws.freeze_panes = ws.cell(row=2, column=len(INFO_OUT) + 1)

    # Column widths — sample header only for speed (many date columns)
    for ci, h in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 2, 30)

    log(f"  Sheet '{acct}' — {len(pivot)} items × {len(week_cols)} weeks")

wb.save(OUT_XLSX)
size_kb = os.path.getsize(OUT_XLSX) / 1024
log(f"Saved: {OUT_XLSX}  ({size_kb:,.1f} KB)  ({time.time() - t0:.1f}s)")

log("")
log("─" * 60)
log("  DONE")
log(f"  Parquet : {OUT_PARQ}")
log(f"  Excel   : {OUT_XLSX}")
log(f"  Rows    : {len(baseline):,}")
log(f"  Accounts: {baseline['ACCOUNT'].nunique()}")
log(f"  Items   : {baseline['ITEM_ID'].nunique():,}")
log("─" * 60)
