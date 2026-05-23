"""
15_seasonality.py  (v2 — Product-Family-level seasonality)

Computes seasonal indices aggregated at PRODUCT FAMILY level across all
accounts and items, then applies them per-row in the item-level parquet.

Architecture
------------
v1: SI per account × item   → noisy, sparse
v2: SI per PRODFAM          → stable, clean signal

Calculation flow
----------------
1. Load all account sheets and combine
2. Group by PRODFAM → sum weekly sales → one SI profile per family
3. Compute item-level baseline (trimmed mean) per account × item
4. Build item × account × week parquet via melt + merge
   - BASELINE_DEMAND  = per account × item trimmed mean
   - SEASONALITY_INDEX = PRODFAM SI for that ISO week
   - SEASONAL_BASELINE = item_baseline × prodfam_SI

Outputs
-------
output/top20_accounts_seasonality.xlsx       compatible format for scripts 16 & 17
output/top20_prodfam_seasonality.xlsx        PRODFAM classification summary (single sheet)
output/prodfam_seasonality.parquet
output/seasonality_demand.parquet
frontend/data_source/prodfam_seasonality.parquet   (Streamlit copy)
frontend/data_source/seasonality_demand.parquet    (Streamlit copy)
"""

import os, re, time, datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE         = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE       = os.path.join(BASE, "output", "top20_accounts_weekly_clean.xlsx")
FE_DIR  = os.path.join(BASE, "frontend", "data_source")

OUT_COMPAT  = os.path.join(BASE, "output", "top20_accounts_seasonality.xlsx")
OUT_PF_XL   = os.path.join(BASE, "output", "top20_prodfam_seasonality.xlsx")
OUT_PF_PQ   = os.path.join(BASE, "output", "prodfam_seasonality.parquet")
OUT_ITEM_PQ = os.path.join(BASE, "output", "seasonality_demand.parquet")

INFO_COLS = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]
Q10, Q90  = 0.10, 0.90

# SI_RANGE = max(SI) / min(SI) thresholds
CLASSES = [
    (1.30, "Non-Seasonal"),
    (1.80, "Mild Seasonal"),
    (2.50, "Moderate Seasonal"),
    (9999, "Strong Seasonal"),
]
CLASS_COLORS = {
    "Strong Seasonal":   "FF0000",
    "Moderate Seasonal": "FFC000",
    "Mild Seasonal":     "92D050",
    "Non-Seasonal":      "D9D9D9",
}
QUARTER_NAMES = {
    1: "Winter (Q1)", 2: "Spring (Q2)",
    3: "Summer (Q3)", 4: "Holiday (Q4)",
}


# ── Utilities ─────────────────────────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}")

def section(title: str):
    log("─" * 60)
    log(f"  {title}")
    log("─" * 60)

def week_num(label: str) -> int:
    m = re.search(r"W(\d+)", str(label))
    return int(m.group(1)) if m else 0

def trimmed_mean(arr: np.ndarray) -> float:
    """Q10–Q90 clipped mean of non-zero values."""
    selling = arr[arr > 0]
    if len(selling) == 0:
        return 0.0
    lo = np.quantile(selling, Q10)
    hi = np.quantile(selling, Q90)
    return round(float(np.clip(selling, lo, hi).mean()), 2)

def compute_prodfam_si(totals: np.ndarray, week_nums: np.ndarray) -> tuple:
    """
    totals    : (n_weeks,) sum of all sales across all items+accounts in this PRODFAM
    week_nums : (n_weeks,) ISO week number (1–52) for each column
    Returns   : (si array shape (52,), baseline float)
    """
    baseline = trimmed_mean(totals)
    si = np.zeros(52, dtype=float)
    if baseline == 0:
        return si, 0.0
    for w in range(1, 53):
        mask = (week_nums == w)
        if mask.any():
            si[w - 1] = round(float(totals[mask].mean() / baseline), 4)
    return si, baseline

def classify_si(si: np.ndarray) -> tuple:
    """Returns (class_label, si_range, peak_weeks_str, peak_season_str)."""
    nz = si[si > 0]
    if len(nz) == 0:
        return "Non-Seasonal", 1.0, "—", "—"
    si_range = round(float(nz.max() / nz.min()), 3)
    label = next(cls for thresh, cls in CLASSES if si_range < thresh)
    peak_weeks = [f"W{w+1:02d}" for w, v in enumerate(si) if v >= 1.10]
    q_avg = {q: si[(q-1)*13 : q*13].mean() for q in range(1, 5)}
    peak_q = max(q_avg, key=q_avg.get)
    return label, si_range, ",".join(peak_weeks[:8]) or "—", QUARTER_NAMES[peak_q]


# ── 1. Load all account sheets ─────────────────────────────────────────────────

section("Load Source File")
t_start = time.time()
xl = pd.ExcelFile(SOURCE)
log(f"Sheets ({len(xl.sheet_names)}): {xl.sheet_names}")

frames = []
for sheet in xl.sheet_names:
    df = xl.parse(sheet)
    avail = [c for c in INFO_COLS if c in df.columns]
    wc    = sorted([c for c in df.columns if c not in avail and c != "GRAND_TOTAL"
                    and isinstance(c, str) and c[0].isdigit() and "W" in c])
    sub = df[avail + wc].copy()
    sub["ACCOUNT"] = sheet
    frames.append(sub)

combined = pd.concat(frames, ignore_index=True)
combined["PRODFAM"] = (combined["PRODFAM"].fillna("").astype(str)
                       .str.strip().replace("", "Unknown"))
combined["ITEM_ID"] = pd.to_numeric(combined["ITEM_ID"], errors="coerce").fillna(0).astype(int)

week_cols   = sorted([c for c in combined.columns
                      if isinstance(c, str) and c[0].isdigit() and "W" in c
                      and c not in INFO_COLS + ["ACCOUNT"]])
week_nums_a = np.array([week_num(c) for c in week_cols])

log(f"Combined rows : {len(combined):,}  |  Items : {combined['ITEM_ID'].nunique():,}")
log(f"Families      : {sorted(combined['PRODFAM'].unique())}")
log(f"Week columns  : {len(week_cols)}  ({week_cols[0]} → {week_cols[-1]})")


# ── 2. Compute PRODFAM-level SI profiles ──────────────────────────────────────

section("Compute Product-Family Seasonality")

prodfams    = sorted(combined["PRODFAM"].unique())
si_lookup   = {}    # (PRODFAM, week_num) → float
pf_baseline = {}    # PRODFAM → float
pf_meta     = []

for pf in prodfams:
    mask   = combined["PRODFAM"] == pf
    totals = combined.loc[mask, week_cols].to_numpy(dtype=float).sum(axis=0)
    si, base = compute_prodfam_si(totals, week_nums_a)
    cls, si_range, peak_wks, peak_season = classify_si(si)

    pf_baseline[pf] = base
    for w in range(1, 53):
        si_lookup[(pf, w)] = si[w - 1]

    row = {
        "PRODFAM":           pf,
        "TOTAL_ITEMS":       int(combined.loc[mask, "ITEM_ID"].nunique()),
        "BASELINE_WEEKLY":   base,
        "PEAK_SI":           round(float(si.max()), 4),
        "TROUGH_SI":         round(float(si[si > 0].min()), 4) if (si > 0).any() else 0.0,
        "SI_RANGE":          si_range,
        "PEAK_WEEKS":        peak_wks,
        "PEAK_SEASON":       peak_season,
        "SEASONALITY_CLASS": cls,
    }
    for w in range(1, 53):
        row[f"SI_W{w:02d}"] = si[w - 1]
    pf_meta.append(row)

    log(f"  {pf:<20} {cls:<18} SI_range={si_range:.2f}  baseline={base:,.0f}")

prodfam_df = pd.DataFrame(pf_meta)


# ── 3. Compute per account×item baselines (vectorized) ────────────────────────

section("Compute Item Baselines")

qty_matrix = combined[week_cols].to_numpy(dtype=float)
baselines  = np.array([trimmed_mean(qty_matrix[i]) for i in range(len(combined))])
combined["BASELINE_DEMAND"] = baselines
log(f"Item baselines: min={baselines[baselines>0].min():.1f}  max={baselines.max():.1f}")


# ── 4. Build item × account × week parquet via melt + merge ───────────────────

section("Build Item Parquet")

id_vars  = [c for c in INFO_COLS + ["ACCOUNT", "BASELINE_DEMAND"] if c in combined.columns]
long_df  = combined[id_vars + week_cols].melt(
    id_vars=id_vars, value_vars=week_cols,
    var_name="WEEK_LABEL", value_name="ACTUAL_WEEKLY_DEMAND",
)
long_df["WEEK_NUM"] = long_df["WEEK_LABEL"].map(week_num)
long_df["YEAR"]     = long_df["WEEK_LABEL"].str[:4].astype(int)
long_df["ACCOUNT"]  = long_df["ACCOUNT"].str.title()

# Merge PRODFAM SI values
si_df = pd.DataFrame(
    [{"PRODFAM": pf, "WEEK_NUM": w, "SEASONALITY_INDEX": si_lookup[(pf, w)]}
     for pf, w in si_lookup],
)
long_df = long_df.merge(si_df, on=["PRODFAM", "WEEK_NUM"], how="left")
long_df["SEASONALITY_INDEX"] = long_df["SEASONALITY_INDEX"].fillna(0.0)
long_df["SEASONAL_BASELINE"] = (long_df["BASELINE_DEMAND"] * long_df["SEASONALITY_INDEX"]).round(2)

# Type cleanup
long_df["ITEM_ID"]              = long_df["ITEM_ID"].astype(int)
long_df["ACTUAL_WEEKLY_DEMAND"] = long_df["ACTUAL_WEEKLY_DEMAND"].astype(int)
for col in ["ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]:
    long_df[col] = long_df[col].astype(str).str.strip().replace("nan", "")

col_order = [
    "ACCOUNT", "ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM",
    "YEAR", "WEEK_NUM", "WEEK_LABEL",
    "ACTUAL_WEEKLY_DEMAND", "BASELINE_DEMAND", "SEASONALITY_INDEX", "SEASONAL_BASELINE",
]
item_df = (long_df[col_order]
           .sort_values(["ACCOUNT", "ITEM_ID", "YEAR", "WEEK_NUM"])
           .reset_index(drop=True))

log(f"Rows: {len(item_df):,}  |  Non-zero: {(item_df['ACTUAL_WEEKLY_DEMAND'] > 0).sum():,}")
log(f"Accounts: {item_df['ACCOUNT'].nunique()}  |  Items: {item_df['ITEM_ID'].nunique():,}")


# ── 5. Build PRODFAM × WEEK long-format parquet ───────────────────────────────

section("Build PRODFAM Parquet")

pf_long = pd.DataFrame([
    {
        "PRODFAM":           pf,
        "WEEK_NUM":          w,
        "SEASONALITY_INDEX": si_lookup[(pf, w)],
        "BASELINE_WEEKLY":   pf_baseline[pf],
        "SEASONAL_BASELINE": round(pf_baseline[pf] * si_lookup[(pf, w)], 2),
    }
    for pf in prodfams
    for w in range(1, 53)
])
log(f"PRODFAM parquet rows: {len(pf_long)}")


# ── 6. Save parquets ──────────────────────────────────────────────────────────

section("Save Parquets")

os.makedirs(FE_DIR, exist_ok=True)

for path in [OUT_PF_PQ, os.path.join(FE_DIR, "prodfam_seasonality.parquet")]:
    pf_long.to_parquet(path, index=False)
for path in [OUT_ITEM_PQ, os.path.join(FE_DIR, "seasonality_demand.parquet")]:
    item_df.to_parquet(path, index=False)

log(f"Saved: {OUT_PF_PQ}")
log(f"Saved: {OUT_ITEM_PQ}")
log(f"Copied to frontend/data_source/")


# ── 7. Save PRODFAM classification Excel ──────────────────────────────────────

section("Save PRODFAM Excel")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SI_FILL     = PatternFill("solid", fgColor="FFF2CC")
CENTER      = Alignment(horizontal="center")

wb = Workbook()
ws = wb.active
ws.title = "PRODFAM_SEASONALITY"

summary_cols = ["PRODFAM", "TOTAL_ITEMS", "BASELINE_WEEKLY", "PEAK_SI",
                "TROUGH_SI", "SI_RANGE", "PEAK_WEEKS", "PEAK_SEASON", "SEASONALITY_CLASS"]
si_cols      = [f"SI_W{w:02d}" for w in range(1, 53)]
all_cols     = summary_cols + si_cols
cls_idx      = all_cols.index("SEASONALITY_CLASS") + 1

for ci, h in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER

for ri, row_data in enumerate(prodfam_df[all_cols].itertuples(index=False), 2):
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if all_cols[ci - 1].startswith("SI_W"):
            cell.fill = SI_FILL
        if ci == cls_idx:
            color = CLASS_COLORS.get(str(val), "FFFFFF")
            cell.fill   = PatternFill("solid", fgColor=color)
            cell.font   = Font(bold=True, color="000000" if val == "Non-Seasonal" else "FFFFFF")
            cell.alignment = CENTER

for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 28)

wb.save(OUT_PF_XL)
log(f"Saved: {OUT_PF_XL}")


# ── 8. Save backward-compat Excel for scripts 16 & 17 ────────────────────────

section("Save Compatibility Excel (for scripts 16 & 17)")

compat_wb = Workbook()
compat_wb.remove(compat_wb.active)

for sheet in xl.sheet_names:
    df_src = xl.parse(sheet)
    avail  = [c for c in INFO_COLS if c in df_src.columns]
    wc     = sorted([c for c in df_src.columns if c not in avail and c != "GRAND_TOTAL"
                     and isinstance(c, str) and c[0].isdigit() and "W" in c])
    out    = df_src[avail + wc].sort_values("ITEM_ID").reset_index(drop=True)

    ws_c = compat_wb.create_sheet(title=sheet[:31])
    all_c = avail + wc
    for ci, h in enumerate(all_c, 1):
        cell = ws_c.cell(row=1, column=ci, value=str(h))
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for ri, row_data in enumerate(out.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            ws_c.cell(row=ri, column=ci, value=val)
    for col in ws_c.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws_c.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

compat_wb.save(OUT_COMPAT)
log(f"Saved compat Excel: {OUT_COMPAT}")


# ── Summary ───────────────────────────────────────────────────────────────────

section("DONE")
log(f"Total runtime : {time.time() - t_start:.1f}s")
log(f"Families      : {len(prodfams)}")
log(f"Seasonality   : {prodfam_df['SEASONALITY_CLASS'].value_counts().to_dict()}")
log(f"Item rows     : {len(item_df):,}")
log("")
log(f"  Excel  : {OUT_PF_XL}")
log(f"  Parquet: {OUT_ITEM_PQ}")
log(f"  Parquet: {OUT_PF_PQ}")
