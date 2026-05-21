"""
14_baseline_from_weekly.py

Input  : top20_accounts_weekly.xlsx  (items × week columns, one sheet per account)
Output : top20_accounts_baseline.xlsx (same structure + BASELINE_DEMAND column)

Baseline method — trimmed mean (retail forecasting standard):
  1. For each item, collect all weekly demand values across all weeks
  2. Keep only non-zero weeks (selling weeks) for outlier detection
  3. SPIKE_CAP  = 90th percentile of non-zero weeks
     (weeks above this are promo spikes / bulk loads)
  4. DIP_FLOOR  = 10th percentile of non-zero weeks
     (weeks below this are abnormal dips / stock-outs)
  5. CAPPED_QTY = clip(actual, DIP_FLOOR, SPIKE_CAP) for every non-zero week
  6. BASELINE   = mean(CAPPED_QTY across all non-zero weeks)
     → this is the "normal" weekly demand, spike-and-dip removed

Zero weeks (no orders) are excluded from baseline calculation but preserved
in the output as-is — they represent genuine off-weeks.

Output structure (same as source):
  ITEM_ID | ITEM_DESCR | BRAND | VENDOR | PRODFAM | 2023-W01 … 2026-W13 | GRAND_TOTAL | BASELINE_DEMAND
  One sheet per account, 20 sheets total.

Run from Dataprocessing/:
    python3 14_baseline_from_weekly.py
"""

import os
import time
import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE   = os.path.join(BASE, "output", "top20_accounts_weekly.xlsx")
OUT_XLSX = os.path.join(BASE, "output", "top20_accounts_baseline.xlsx")

INFO_COLS = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

SPIKE_PCT = 0.90   # weeks above this percentile → promo spike
DIP_PCT   = 0.10   # weeks below this percentile → abnormal dip

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue header
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")   # light blue — GRAND_TOTAL
BASELINE_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green — BASELINE_DEMAND
ZERO_FILL     = PatternFill("solid", fgColor="F2F2F2")   # grey — zero cells
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT     = Font(bold=True, size=10)
CENTER        = Alignment(horizontal="center")


def log(msg: str, level: str = "INFO"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [{level}] {msg}")


def section(title: str):
    log("─" * 60)
    log(f"  STEP: {title}")
    log("─" * 60)


def compute_baseline(row_values: np.ndarray) -> float:
    """
    Trimmed mean baseline for one item's weekly demand vector.

    row_values : array of weekly quantities (may include zeros)

    Steps:
      1. Isolate non-zero selling weeks
      2. Compute spike cap (Q90) and dip floor (Q10) from those weeks
      3. Clip each selling week to [dip_floor, spike_cap]
      4. Return mean of clipped selling weeks  (→ "normal" weekly demand)

    Returns 0.0 if item has no selling weeks at all.
    """
    selling = row_values[row_values > 0]
    if len(selling) == 0:
        return 0.0

    spike_cap = np.quantile(selling, SPIKE_PCT)
    dip_floor = np.quantile(selling, DIP_PCT)

    # If all weeks are the same value both quantiles equal → no clipping needed
    clipped = np.clip(selling, dip_floor, spike_cap)
    return round(float(clipped.mean()), 2)


section("Load Source File")
xl       = pd.ExcelFile(SOURCE)
log(f"Source : {SOURCE}")
log(f"Sheets : {xl.sheet_names}")

section("Process Each Sheet & Write Output")
t0_total = time.time()

wb = openpyxl.Workbook()
wb.remove(wb.active)

total_spikes = 0
total_dips   = 0

for sheet in xl.sheet_names:
    t0      = time.time()
    df      = xl.parse(sheet)

    avail_info = [c for c in INFO_COLS if c in df.columns]
    week_cols  = sorted([c for c in df.columns if c not in avail_info and c != "GRAND_TOTAL"])

    qty_matrix = df[week_cols].to_numpy(dtype=float)   # shape: (n_items, n_weeks)

    # ── Compute baseline per item ─────────────────────────────────────────────
    baselines  = np.array([compute_baseline(qty_matrix[i]) for i in range(len(df))])

    # ── Diagnostics: count capped/floored weeks per sheet ────────────────────
    for i in range(len(df)):
        selling = qty_matrix[i][qty_matrix[i] > 0]
        if len(selling) == 0:
            continue
        cap   = np.quantile(selling, SPIKE_PCT)
        floor = np.quantile(selling, DIP_PCT)
        total_spikes += int((selling > cap).sum())
        total_dips   += int((selling < floor).sum())

    # ── Rebuild output DataFrame ──────────────────────────────────────────────
    out            = df[avail_info + week_cols].copy()
    out["GRAND_TOTAL"]      = out[week_cols].sum(axis=1)
    out["BASELINE_DEMAND"]  = baselines
    out = out.sort_values("GRAND_TOTAL", ascending=False).reset_index(drop=True)

    all_cols        = avail_info + week_cols + ["GRAND_TOTAL", "BASELINE_DEMAND"]
    grand_col_idx   = len(avail_info) + len(week_cols) + 1   # 1-based
    baseline_col_idx= len(all_cols)                           # 1-based

    # ── Write sheet ───────────────────────────────────────────────────────────
    ws = wb.create_sheet(title=sheet[:31])

    # Header
    for ci, h in enumerate(all_cols, 1):
        cell           = ws.cell(row=1, column=ci, value=str(h))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # Data rows
    for ri, row in enumerate(out[all_cols].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == baseline_col_idx:
                cell.fill = BASELINE_FILL
                cell.font = BOLD_FONT
            elif ci == grand_col_idx:
                cell.fill = TOTAL_FILL
                cell.font = BOLD_FONT
            elif ci > len(avail_info) and (val == 0 or val == "0"):
                cell.fill = ZERO_FILL

    # Freeze panes after info columns
    ws.freeze_panes = ws.cell(row=2, column=len(avail_info) + 1)

    # Column widths
    for ci, h in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 2, 30)

    spike_count = sum(
        int((qty_matrix[i][qty_matrix[i] > 0] > np.quantile(qty_matrix[i][qty_matrix[i] > 0], SPIKE_PCT)).sum())
        for i in range(len(df))
        if (qty_matrix[i] > 0).any()
    )

    log(f"  {sheet:<30} {len(out):>5} items | baseline range: "
        f"{baselines[baselines>0].min() if (baselines>0).any() else 0:.1f}"
        f" – {baselines.max():.1f}  ({time.time()-t0:.1f}s)")

section("Save Workbook")
t0 = time.time()
wb.save(OUT_XLSX)
size_kb = os.path.getsize(OUT_XLSX) / 1024
log(f"Saved : {OUT_XLSX}")
log(f"Size  : {size_kb:,.1f} KB")
log(f"Spike weeks detected & capped : {total_spikes:,}")
log(f"Dip   weeks detected & floored: {total_dips:,}")
log(f"Total runtime: {time.time()-t0_total:.1f}s")
