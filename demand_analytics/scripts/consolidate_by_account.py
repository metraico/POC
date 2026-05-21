"""
consolidate_by_account.py

Two modes:
  1. Single account  → python3 consolidate_by_account.py "total wine"
  2. Top-20 accounts → python3 consolidate_by_account.py --top20
     Produces one Excel file with 20 sheets, one per account (by customer count).

No args → lists all available account names.
"""

import os
import sys
import re
import time
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))           # demand_analytics/scripts/
_DA_DIR      = os.path.dirname(_SCRIPTS_DIR)                         # demand_analytics/
BASE         = os.path.dirname(_DA_DIR)                              # Dataprocessing/
ACCOUNT_FILE = os.path.join(BASE, "OneDrive_1_08-05-2026", "CustomerNames_with_AccountNames.xlsx")
DEMAND_FILE  = os.path.join(BASE, "OneDrive_1_08-05-2026", "DEMAND ORDER LINES.csv")
ITEM_MASTER  = os.path.join(BASE, "OneDrive_1_08-05-2026", "ITEM MASTER.xlsx")
OUTPUT_DIR   = os.path.join(_DA_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

LOG_LINES = []

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
ZERO_FILL   = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FONT  = Font(bold=True, size=10)
CENTER      = Alignment(horizontal="center")


def log(msg: str, *, level: str = "INFO"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] [{level}] {msg}"
    print(line)
    LOG_LINES.append(line)


def section(title: str):
    divider = "─" * 60
    log(divider)
    log(f"  STEP: {title}")
    log(divider)


def save_log(slug: str):
    log_path = os.path.join(OUTPUT_DIR, f"{slug}_run.log")
    with open(log_path, "w") as f:
        f.write("\n".join(LOG_LINES))
    print(f"\n[LOG] Full log written → {log_path}")


# ── Step 1: Load account mapping ──────────────────────────────────────────────
def load_account_mapping() -> pd.DataFrame:
    section("Load Account Mapping")
    t0 = time.time()
    df = pd.read_excel(ACCOUNT_FILE)
    df.columns = df.columns.str.strip()

    col_map = {}
    for c in df.columns:
        if "customer" in c.lower():
            col_map[c] = "Customer Name"
        elif "account" in c.lower():
            col_map[c] = "Account Name"
    df = df.rename(columns=col_map)

    df["Account Name"]  = df["Account Name"].astype(str).str.strip().str.lower()
    df["Customer Name"] = df["Customer Name"].astype(str).str.strip()

    log(f"Loaded {len(df):,} rows — {df['Account Name'].nunique()} unique accounts")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return df


# ── Step 2: Get top-N accounts by customer count ──────────────────────────────
def get_top_accounts(mapping_df: pd.DataFrame, n: int = 20) -> list[str]:
    section(f"Resolve Top {n} Accounts by Customer Count")
    counts = (
        mapping_df[~mapping_df["Account Name"].isin(["grand total", "nan"])]
        .groupby("Account Name")["Customer Name"]
        .count()
        .sort_values(ascending=False)
        .head(n)
    )
    for rank, (name, cnt) in enumerate(counts.items(), 1):
        log(f"  {rank:>2}. {name:<35} {cnt} customers")
    return counts.index.tolist()


# ── Step 3: Resolve customers for one account ─────────────────────────────────
def resolve_customers(mapping_df: pd.DataFrame, account_name: str) -> tuple[list[str], list[int]]:
    norm   = account_name.strip().lower()
    subset = mapping_df[mapping_df["Account Name"] == norm]

    if subset.empty:
        log(f"Account '{account_name}' not found.", level="ERROR")
        return [], []

    customer_names = subset["Customer Name"].tolist()
    customer_ids   = []
    for name in customer_names:
        m = re.search(r"/\s*(\d+)\s*$", name)
        if m:
            customer_ids.append(int(m.group(1)))
        else:
            log(f"  WARNING: could not extract ID from '{name}'", level="WARN")

    log(f"  '{account_name}' → {len(customer_names)} customers, {len(customer_ids)} IDs extracted")
    return customer_names, customer_ids


# ── Step 4: Load full DEMAND ORDER LINES once ─────────────────────────────────
def load_full_demand() -> pd.DataFrame:
    section("Load DEMAND ORDER LINES (full)")
    t0 = time.time()
    log(f"Reading {DEMAND_FILE} …")
    demand = pd.read_csv(
        DEMAND_FILE,
        usecols=["ITEM", "CUSTOMERID", "ITEMNAME", "YEAR", "MONTH", "DAY", "QUANTITY"],
        dtype={"CUSTOMERID": int, "ITEM": object},
    )
    log(f"Total rows: {len(demand):,}")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return demand


# ── Step 5: Filter demand for a set of customer IDs ───────────────────────────
def filter_demand(demand: pd.DataFrame, customer_ids: list[int], account_name: str) -> pd.DataFrame:
    section(f"Filter Demand — {account_name}")
    t0 = time.time()

    filtered = demand[demand["CUSTOMERID"].isin(customer_ids)].copy()
    log(f"Rows matched: {len(filtered):,}")

    before   = len(filtered)
    filtered = filtered[filtered["QUANTITY"] > 0].copy()
    log(f"After dropping zero/negative: {len(filtered):,}  (removed {before - len(filtered):,})")

    if filtered.empty:
        log(f"No positive-quantity rows for '{account_name}'", level="WARN")
        return filtered

    filtered["DATE"] = pd.to_datetime(
        filtered[["YEAR", "MONTH", "DAY"]].rename(
            columns={"YEAR": "year", "MONTH": "month", "DAY": "day"}
        )
    ).dt.date

    # Build ITEM → ITEMNAME fallback map for items missing from ITEM MASTER
    filtered["_ITEMNAME"] = filtered["ITEMNAME"].astype(str).str.strip()

    cust_summary = filtered.groupby("CUSTOMERID").agg(
        rows=("QUANTITY", "count"), units=("QUANTITY", "sum")
    )
    for cid, row in cust_summary.iterrows():
        log(f"  CUSTOMERID {cid:>8} — {row['rows']:>6,} rows, {row['units']:>8,.0f} units")

    log(f"Date range: {filtered['DATE'].min()} → {filtered['DATE'].max()}")
    log(f"Unique items: {filtered['ITEM'].nunique():,}")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return filtered


# ── Step 6: Load ITEM MASTER ──────────────────────────────────────────────────
def load_item_reference() -> pd.DataFrame:
    section("Load Item Reference (ITEM MASTER)")
    t0 = time.time()
    master = pd.read_excel(
        ITEM_MASTER,
        usecols=["ITEM", "DESCR", "BRAND", "VENDOR", "PRODFAM"],
    )
    master["ITEM"] = pd.to_numeric(master["ITEM"], errors="coerce")
    master = master.rename(columns={"ITEM": "ITEM_ID", "DESCR": "ITEM_DESCR"})
    log(f"ITEM MASTER rows: {len(master):,}")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return master


# ── Step 7: Aggregate & pivot ─────────────────────────────────────────────────
def build_pivot(demand_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    section("Aggregate & Pivot")
    t0 = time.time()

    demand_df["ITEM"] = demand_df["ITEM"].astype(str).str.strip()
    # Build fallback name map before aggregation
    itemname_map = (
        demand_df.groupby("ITEM")["_ITEMNAME"]
        .first()
        .to_dict()
    )
    agg = demand_df.groupby(["ITEM", "DATE"])["QUANTITY"].sum().reset_index()
    log(f"Aggregated rows (item × date): {len(agg):,}")

    pivot = agg.pivot_table(
        index="ITEM", columns="DATE", values="QUANTITY", aggfunc="sum", fill_value=0
    ).reset_index()
    pivot.columns.name = None

    date_cols = [c for c in pivot.columns if c != "ITEM"]
    pivot["GRAND_TOTAL"] = pivot[date_cols].sum(axis=1)
    pivot = pivot.sort_values("GRAND_TOTAL", ascending=False).reset_index(drop=True)

    log(f"Pivot: {len(pivot)} items × {len(date_cols)} dates")
    log(f"Grand total units: {pivot['GRAND_TOTAL'].sum():,.0f}")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return pivot, date_cols, itemname_map


# ── Step 8: Merge item metadata ───────────────────────────────────────────────
def merge_metadata(pivot: pd.DataFrame, item_ref: pd.DataFrame, date_cols: list,
                   itemname_map: dict) -> tuple[pd.DataFrame, pd.DataFrame, list]:
    section("Merge Item Metadata")
    t0 = time.time()

    pivot["ITEM"] = pd.to_numeric(pivot["ITEM"], errors="coerce")
    pivot = pivot.rename(columns={"ITEM": "ITEM_ID"})
    merged = pivot.merge(item_ref, on="ITEM_ID", how="left")

    info_cols      = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]
    available_info = [c for c in info_cols if c in merged.columns]

    # Split matched vs unmatched
    matched_df   = merged[merged["ITEM_DESCR"].notna()].copy()
    unmatched_df = merged[merged["ITEM_DESCR"].isna()].copy()

    # For unmatched: use ITEMNAME from demand CSV as fallback description
    if not unmatched_df.empty:
        unmatched_df["ITEM_DESCR"] = unmatched_df["ITEM_ID"].astype(str).map(
            {str(k): v for k, v in itemname_map.items()}
        )
        unmatched_df["BRAND"]   = "NOT IN ITEM MASTER"
        unmatched_df["VENDOR"]  = "NOT IN ITEM MASTER"
        unmatched_df["PRODFAM"] = "NOT IN ITEM MASTER"
        missing_ids = unmatched_df["ITEM_ID"].tolist()
        log(f"  Unmatched IDs: {missing_ids[:10]}{'...' if len(missing_ids) > 10 else ''}", level="WARN")

    matched_df   = matched_df[available_info + date_cols + ["GRAND_TOTAL"]]
    unmatched_df = unmatched_df[available_info + date_cols + ["GRAND_TOTAL"]]

    log(f"Matched to ITEM MASTER : {len(matched_df)}")
    log(f"NOT in ITEM MASTER     : {len(unmatched_df)}")
    log(f"Step duration: {time.time() - t0:.2f}s")
    return matched_df, unmatched_df, available_info


# ── Write one sheet into an existing workbook ─────────────────────────────────
def write_sheet(wb: openpyxl.Workbook, final: pd.DataFrame, info_cols: list,
                date_cols: list, sheet_name: str):
    ws = wb.create_sheet(title=sheet_name[:31])
    headers       = final.columns.tolist()
    grand_col_idx = headers.index("GRAND_TOTAL") + 1

    for col_idx, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col_idx, value=str(h))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    for row_idx, row in enumerate(final.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == grand_col_idx:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            elif col_idx > len(info_cols) and (val == 0 or val == "0"):
                cell.fill = ZERO_FILL

    for col_idx, h in enumerate(headers, 1):
        col_vals = [str(h)] + [
            str(ws.cell(row=r, column=col_idx).value or "")
            for r in range(2, min(len(final) + 2, 52))
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(len(v) for v in col_vals) + 2, 30
        )

    ws.freeze_panes = ws.cell(row=2, column=len(info_cols) + 1)


# ── Run pipeline for one account, return (final_df, info_cols, date_cols) ────
def process_account(account_name: str, mapping_df: pd.DataFrame,
                    demand_full: pd.DataFrame, item_ref: pd.DataFrame):
    log(f"\n{'='*60}")
    log(f"  PROCESSING: {account_name.upper()}")
    log(f"{'='*60}")

    customer_names, customer_ids = resolve_customers(mapping_df, account_name)
    if not customer_ids:
        return None, None, None, None, customer_names, customer_ids

    demand_df            = filter_demand(demand_full, customer_ids, account_name)
    if demand_df.empty:
        return None, None, None, None, customer_names, customer_ids

    pivot, date_cols, itemname_map          = build_pivot(demand_df)
    matched_df, unmatched_df, info_cols     = merge_metadata(pivot, item_ref, date_cols, itemname_map)
    return matched_df, unmatched_df, info_cols, date_cols, customer_names, customer_ids


# ── Final summary ─────────────────────────────────────────────────────────────
def final_summary_multi(results: list, output_path: str, total_time: float):
    section("FINAL SUMMARY — TOP 20 ACCOUNTS")
    for r in results:
        status = "OK" if r["ok"] else "SKIPPED"
        log(f"  [{status:^7}] {r['account']:<35} "
            f"{r['customers']:>3} customers  "
            f"{r.get('items', 0):>5} items  "
            f"{r.get('units', 0):>10,.0f} units")
    log(f"\nOutput file : {output_path}")
    log(f"Total runtime: {total_time:.1f}s")
    log(f"Completed at : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


def final_summary_single(account_name, customer_names, customer_ids,
                         demand_df, final_df, date_cols, output_path):
    section("FINAL SUMMARY")
    log(f"Account Name    : {account_name}")
    log(f"Customer Names  : {len(customer_names)}")
    log(f"Customer IDs    : {customer_ids}")
    log(f"Demand rows     : {len(demand_df):,}")
    log(f"Total units     : {demand_df['QUANTITY'].sum():,.0f}")
    log(f"Unique items    : {final_df['ITEM_ID'].nunique()}")
    log(f"Date range      : {min(date_cols)} → {max(date_cols)}")
    log(f"Output file     : {output_path}")
    log(f"Completed at    : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    run_start = time.time()

    # ── Mode: no args → list accounts ────────────────────────────────────────
    if len(sys.argv) < 2:
        print("\nUsage:")
        print("  python3 consolidate_by_account.py \"<account name>\"   # single account")
        print("  python3 consolidate_by_account.py --top20             # top 20 accounts")
        print("\nAvailable account names:")
        mapping = pd.read_excel(ACCOUNT_FILE)
        mapping.columns = mapping.columns.str.strip()
        acct_col = next(c for c in mapping.columns if "account" in c.lower())
        counts = mapping.groupby(acct_col)[mapping.columns[0]].count().sort_values(ascending=False)
        for name, cnt in counts.items():
            if name.lower() not in ("grand total", "nan"):
                print(f"  • {name}  ({cnt})")
        sys.exit(0)

    # ── Mode: --top20 ─────────────────────────────────────────────────────────
    if sys.argv[1].strip() == "--top20":
        output_path = os.path.join(OUTPUT_DIR, "top20_accounts_consolidated.xlsx")
        log(f"TOP-20 mode — output: {output_path}")

        mapping_df  = load_account_mapping()
        top20       = get_top_accounts(mapping_df, n=20)
        demand_full = load_full_demand()
        item_ref    = load_item_reference()

        missing_path = os.path.join(OUTPUT_DIR, "top20_accounts_not_in_master.xlsx")

        wb         = openpyxl.Workbook()
        wb.remove(wb.active)
        wb_missing = openpyxl.Workbook()
        wb_missing.remove(wb_missing.active)
        results        = []
        any_missing    = False

        for rank, account_name in enumerate(top20, 1):
            log(f"\n[{rank}/20] {account_name}")
            matched_df, unmatched_df, info_cols, date_cols, cust_names, cust_ids = process_account(
                account_name, mapping_df, demand_full, item_ref
            )

            if matched_df is None:
                results.append({"account": account_name, "customers": len(cust_names), "ok": False})
                continue

            write_sheet(wb, matched_df, info_cols, date_cols, account_name)
            if unmatched_df is not None and not unmatched_df.empty:
                write_sheet(wb_missing, unmatched_df, info_cols, date_cols, account_name)
                any_missing = True
                log(f"  {len(unmatched_df)} items not in ITEM MASTER — added to missing file", level="WARN")
            results.append({
                "account":   account_name,
                "customers": len(cust_names),
                "items":     matched_df["ITEM_ID"].nunique(),
                "units":     matched_df["GRAND_TOTAL"].sum(),
                "ok":        True,
            })
            log(f"  Sheet '{account_name}' written — {len(matched_df)} items")

        section("Saving Workbooks")
        t0 = time.time()
        wb.save(output_path)
        size_kb = os.path.getsize(output_path) / 1024
        log(f"Saved: {output_path}  ({size_kb:,.1f} KB)  in {time.time() - t0:.1f}s")

        if any_missing:
            wb_missing.save(missing_path)
            size_kb2 = os.path.getsize(missing_path) / 1024
            log(f"Saved: {missing_path}  ({size_kb2:,.1f} KB)")
        else:
            log("No items missing from ITEM MASTER — missing file not created.")

        final_summary_multi(results, output_path, time.time() - run_start)
        save_log("top20_accounts")

    # ── Mode: single account ──────────────────────────────────────────────────
    else:
        account_name = sys.argv[1].strip()
        account_slug = re.sub(r"[^a-z0-9]+", "_", account_name.lower()).strip("_")
        output_path  = os.path.join(OUTPUT_DIR, f"{account_slug}_consolidated.xlsx")

        log(f"Single-account mode: '{account_name}'")
        log(f"Output: {output_path}")

        mapping_df                              = load_account_mapping()
        customer_names, customer_ids            = resolve_customers(mapping_df, account_name)
        if not customer_ids:
            sys.exit(1)

        demand_full                             = load_full_demand()
        demand_df                               = filter_demand(demand_full, customer_ids, account_name)
        item_ref                                = load_item_reference()
        pivot, date_cols, itemname_map          = build_pivot(demand_df)
        matched_df, unmatched_df, info_cols     = merge_metadata(pivot, item_ref, date_cols, itemname_map)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        write_sheet(wb, matched_df, info_cols, date_cols, account_name)
        if not unmatched_df.empty:
            missing_path = os.path.join(OUTPUT_DIR, f"{account_slug}_not_in_master.xlsx")
            wb_missing   = openpyxl.Workbook()
            wb_missing.remove(wb_missing.active)
            write_sheet(wb_missing, unmatched_df, info_cols, date_cols, account_name)
            wb_missing.save(missing_path)
            log(f"  {len(unmatched_df)} items not in ITEM MASTER → {missing_path}", level="WARN")

        section("Saving Workbook")
        wb.save(output_path)
        size_kb = os.path.getsize(output_path) / 1024
        log(f"Saved: {output_path}  ({size_kb:,.1f} KB)")

        final_summary_single(account_name, customer_names, customer_ids,
                             demand_df, matched_df, date_cols, output_path)
        log(f"Total runtime: {time.time() - run_start:.2f}s")
        save_log(account_slug)


if __name__ == "__main__":
    main()
