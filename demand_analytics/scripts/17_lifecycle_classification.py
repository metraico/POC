import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE  = "../output/top20_accounts_seasonality.xlsx"
OUTPUT_FILE = "../output/top20_accounts_lifecycle.xlsx"

DECLINING_THRESHOLD = -0.30
PARTIAL_YEAR        = "2026"

COLORS = {
    "New":          "4472C4",
    "Evergreen":    "00B050",
    "Declining":    "FF6600",
    "Discontinued": "FF0000",
}


def partial_year_weeks(week_cols, year):
    return sum(1 for w in week_cols if w.startswith(year + "-"))


def classify_lifecycle(totals_by_year, week_cols):
    """Returns {year_str: status} for years with sales history."""
    all_years       = sorted(totals_by_year.keys())
    first_sale_year = next((y for y in all_years if totals_by_year[y] > 0), None)
    if first_sale_year is None:
        return {}

    full_weeks    = 52
    partial_weeks = partial_year_weeks(week_cols, PARTIAL_YEAR)

    result           = {}
    prev_sales       = None
    had_sales_before = False

    for year in all_years:
        sales = totals_by_year[year]
        annualised = (
            sales * (full_weeks / partial_weeks)
            if (year == PARTIAL_YEAR and partial_weeks > 0)
            else sales
        )

        if year < first_sale_year:
            prev_sales = None
            continue

        if year == first_sale_year:
            result[year]     = "New"
            prev_sales       = annualised
            had_sales_before = True
            continue

        if sales == 0:
            if had_sales_before:
                result[year] = "Discontinued"
            prev_sales = annualised
            continue

        yoy = ((annualised - prev_sales) / prev_sales) if (prev_sales and prev_sales > 0) else None
        result[year]     = "Declining" if (yoy is not None and yoy <= DECLINING_THRESHOLD) else "Evergreen"
        prev_sales       = annualised
        had_sales_before = True

    return result


def style_sheet(ws, lc_cols=None):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Color lifecycle status cells
    col_indices = {}
    for i, cell in enumerate(ws[1], 1):
        if cell.value and (str(cell.value).startswith("LC_") or cell.value == "LIFECYCLE_OVERALL"):
            col_indices[i] = cell.value

    for row in ws.iter_rows(min_row=2):
        for col_idx, col_name in col_indices.items():
            cell  = row[col_idx - 1]
            color = COLORS.get(cell.value)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def main():
    xl        = pd.ExcelFile(INPUT_FILE)
    meta_cols = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

    # ── Load all sheets ────────────────────────────────────────────────────────
    frames          = []
    account_item_map = defaultdict(set)   # account → set of ITEM_IDs

    for sheet in xl.sheet_names:
        df       = xl.parse(sheet)
        week_cols_sheet = [
            c for c in df.columns
            if isinstance(c, str) and c[0].isdigit() and "W" in c and not c.startswith("SB")
        ]
        keep = [c for c in meta_cols if c in df.columns] + week_cols_sheet
        frames.append(df[keep])
        for item_id in df["ITEM_ID"].unique():
            account_item_map[sheet].add(item_id)

    all_data  = pd.concat(frames, ignore_index=True)
    week_cols = [
        c for c in all_data.columns
        if isinstance(c, str) and c[0].isdigit() and "W" in c
    ]
    all_years = sorted(set(w.split("-")[0] for w in week_cols))

    # ── Aggregate sales per item across all accounts ───────────────────────────
    meta = all_data[meta_cols].drop_duplicates(subset=["ITEM_ID"])
    agg  = all_data.groupby("ITEM_ID")[week_cols].sum().reset_index()
    agg  = meta.merge(agg, on="ITEM_ID", how="right")

    # ── Classify lifecycle per item (item-level, not per account) ─────────────
    lc_year_cols = [f"LC_{y}" for y in all_years]

    item_lc_rows = []
    for _, item_row in agg.iterrows():
        totals = defaultdict(float)
        for w in week_cols:
            yr = w.split("-")[0]
            totals[yr] += item_row[w]

        lc_map = classify_lifecycle(dict(totals), week_cols)

        row = {c: item_row[c] for c in meta_cols}
        for yr in all_years:
            row[f"LC_{yr}"] = lc_map.get(yr, "No Data")

        # Overall = most recent year that has a real classification
        classified_years = [yr for yr in reversed(all_years) if lc_map.get(yr) not in (None, "No Data")]
        row["LIFECYCLE_OVERALL"] = lc_map[classified_years[0]] if classified_years else "No Data"
        item_lc_rows.append(row)

    item_lc = pd.DataFrame(item_lc_rows)

    # ── Build item × account mapping ──────────────────────────────────────────
    account_rows = []
    for account, item_ids in account_item_map.items():
        for item_id in item_ids:
            account_rows.append({"ACCOUNT": account, "ITEM_ID": item_id})
    account_df = pd.DataFrame(account_rows)

    # ── Final result: one row per item × account ──────────────────────────────
    result = account_df.merge(item_lc, on="ITEM_ID", how="left")
    result = result[["ACCOUNT"] + meta_cols + lc_year_cols + ["LIFECYCLE_OVERALL"]]
    result = result.sort_values(["ACCOUNT", "ITEM_ID"]).reset_index(drop=True)

    # ── Excel output ──────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="LIFECYCLE", index=False)
        style_sheet(writer.sheets["LIFECYCLE"])

        # Counts per year per status (item-level, deduplicated)
        count_rows = []
        for yr in all_years:
            col = f"LC_{yr}"
            vc  = item_lc[col].value_counts()
            for status, cnt in vc.items():
                if status != "No Data":
                    count_rows.append({"YEAR": int(yr), "STATUS": status, "COUNT": cnt})
        counts = pd.DataFrame(count_rows).pivot_table(
            index="YEAR", columns="STATUS", values="COUNT", fill_value=0
        ).reset_index()
        for s in ["New", "Evergreen", "Declining", "Discontinued"]:
            if s not in counts.columns:
                counts[s] = 0
        counts = counts[["YEAR"] + [s for s in ["New", "Evergreen", "Declining", "Discontinued"] if s in counts.columns]]
        counts.to_excel(writer, sheet_name="LIFECYCLE_COUNTS", index=False)
        style_sheet(writer.sheets["LIFECYCLE_COUNTS"])

    print(f"Output saved: {OUTPUT_FILE}")
    print(f"Rows (item × account): {len(result)}")
    print(f"Unique items: {item_lc['ITEM_ID'].nunique()}")
    print(f"\nOverall lifecycle:")
    print(item_lc["LIFECYCLE_OVERALL"].value_counts().to_string())
    print(f"\nBy year:")
    for yr in all_years:
        col = f"LC_{yr}"
        vc  = item_lc[col][item_lc[col] != "No Data"].value_counts()
        print(f"  {yr}: {dict(vc)}")


if __name__ == "__main__":
    main()
