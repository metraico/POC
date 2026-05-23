"""
18_data_quality.py

Removes low-signal items from the weekly source data.

Inclusion rule (item-level, aggregated across ALL accounts and years):
  TOTAL_DEMAND  >= MIN_TOTAL_DEMAND   (at least 500 total units sold)
  AND
  SELLING_WEEKS >= MIN_SELLING_WEEKS  (sold in at least 12 weeks total)

Items failing EITHER condition are excluded from the clean output.

Outputs
-------
output/top20_data_quality.xlsx        all items with quality metrics + flag (audit file)
output/top20_accounts_weekly_clean.xlsx  cleaned source — input for scripts 15/16/17
"""

import os, datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Thresholds (adjust here if needed) ───────────────────────────────────────
MIN_TOTAL_DEMAND  = 500  # minimum total units across ALL accounts + years
MIN_SELLING_WEEKS = 12   # minimum total selling weeks across ALL accounts + years

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE, "output", "top20_accounts_weekly.xlsx")
OUT_QC_XL  = os.path.join(BASE, "output", "top20_data_quality.xlsx")
OUT_CLEAN  = os.path.join(BASE, "output", "top20_accounts_weekly_clean.xlsx")

INFO_COLS  = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
CENTER        = Alignment(horizontal="center")
GHOST_FILL    = PatternFill("solid", fgColor="FF0000")
CLEAN_FILL    = PatternFill("solid", fgColor="00B050")
BOLD_WHITE    = Font(bold=True, color="FFFFFF")


def log(msg):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}")


def style_ws(ws):
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── 1. Load all sheets ────────────────────────────────────────────────────────
log("Loading source data...")
xl = pd.ExcelFile(SOURCE)

all_frames = []
sheet_data = {}   # sheet → original df (for writing clean output)

for sheet in xl.sheet_names:
    df = xl.parse(sheet)
    week_cols = [c for c in df.columns
                 if isinstance(c, str) and c[0].isdigit() and "W" in c
                 and c != "GRAND_TOTAL"]
    df["_TOTAL"]     = df[week_cols].sum(axis=1)
    df["_SELL_WKS"]  = (df[week_cols] > 0).sum(axis=1)
    df["_ACCOUNT"]   = sheet
    sheet_data[sheet] = (df, week_cols)
    keep = [c for c in INFO_COLS if c in df.columns] + ["_TOTAL", "_SELL_WKS", "_ACCOUNT"]
    all_frames.append(df[keep])

combined = pd.concat(all_frames, ignore_index=True)
combined["ITEM_ID"] = pd.to_numeric(combined["ITEM_ID"], errors="coerce").fillna(0).astype(int)


# ── 2. Aggregate to item level ────────────────────────────────────────────────
log("Computing item-level quality metrics...")

item_agg = (
    combined.groupby("ITEM_ID")
    .agg(
        TOTAL_DEMAND      = ("_TOTAL",    "sum"),
        SELLING_WEEKS     = ("_SELL_WKS", "sum"),
        ACCOUNTS_ACTIVE   = ("_ACCOUNT",  lambda x: (x.groupby(x).transform("sum") > 0).nunique()
                             if False else (combined.loc[x.index]
                                            .groupby("_ACCOUNT")["_TOTAL"].sum() > 0).sum()),
        ITEM_DESCR        = ("ITEM_DESCR" if "ITEM_DESCR" in combined.columns else "_ACCOUNT", "first"),
        BRAND             = ("BRAND"      if "BRAND"      in combined.columns else "_ACCOUNT", "first"),
        VENDOR            = ("VENDOR"     if "VENDOR"     in combined.columns else "_ACCOUNT", "first"),
        PRODFAM           = ("PRODFAM"    if "PRODFAM"    in combined.columns else "_ACCOUNT", "first"),
    )
    .reset_index()
)

# Simpler ACCOUNTS_ACTIVE: count accounts that had any sales
accounts_active = (
    combined[combined["_TOTAL"] > 0]
    .groupby("ITEM_ID")["_ACCOUNT"].nunique()
    .reset_index(name="ACCOUNTS_ACTIVE")
)
item_agg = item_agg.drop(columns=["ACCOUNTS_ACTIVE"]).merge(accounts_active, on="ITEM_ID", how="left")
item_agg["ACCOUNTS_ACTIVE"] = item_agg["ACCOUNTS_ACTIVE"].fillna(0).astype(int)
item_agg["AVG_QTY_PER_SELLING_WEEK"] = (
    item_agg["TOTAL_DEMAND"] / item_agg["SELLING_WEEKS"].replace(0, np.nan)
).round(2)

# ── 3. Flag low-signal items (fails EITHER threshold) ────────────────────────
ghost_mask = (
    (item_agg["TOTAL_DEMAND"]  < MIN_TOTAL_DEMAND) |
    (item_agg["SELLING_WEEKS"] < MIN_SELLING_WEEKS)
)
item_agg["DEMAND_QUALITY"] = np.where(ghost_mask, "Insufficient Data", "Analysable")

ghost_ids  = set(item_agg.loc[ghost_mask, "ITEM_ID"])
clean_ids  = set(item_agg.loc[~ghost_mask, "ITEM_ID"])

log(f"Total items   : {len(item_agg):,}")
log(f"Excluded      : {len(ghost_ids):,}  (TOTAL<{MIN_TOTAL_DEMAND} OR SELL_WKS<{MIN_SELLING_WEEKS})")
log(f"Clean items   : {len(clean_ids):,}")


# ── 4. Save data quality Excel ────────────────────────────────────────────────
log("Saving data quality report...")

col_order = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM",
             "TOTAL_DEMAND", "SELLING_WEEKS", "AVG_QTY_PER_SELLING_WEEK",
             "ACCOUNTS_ACTIVE", "DEMAND_QUALITY"]
item_agg  = item_agg[col_order].sort_values("TOTAL_DEMAND").reset_index(drop=True)
ghost_df  = item_agg[item_agg["DEMAND_QUALITY"] == "Insufficient Data"].reset_index(drop=True)
clean_df  = item_agg[item_agg["DEMAND_QUALITY"] == "Analysable"].reset_index(drop=True)

dq_col_idx = col_order.index("DEMAND_QUALITY") + 1

wb = Workbook()
wb.remove(wb.active)

for title, data in [("ALL_ITEMS", item_agg), ("GHOST_SKUS", ghost_df), ("CLEAN_ITEMS", clean_df)]:
    ws = wb.create_sheet(title=title)
    for ci, h in enumerate(col_order, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for ri, row_vals in enumerate(data[col_order].itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == dq_col_idx:
                if val == "Insufficient Data":
                    cell.fill, cell.font = GHOST_FILL, BOLD_WHITE
                else:
                    cell.fill, cell.font = CLEAN_FILL, BOLD_WHITE
                cell.alignment = CENTER
    style_ws(ws)

wb.save(OUT_QC_XL)
log(f"Saved: {OUT_QC_XL}")
log(f"  ALL_ITEMS : {len(item_agg):,} rows")
log(f"  GHOST_SKUS: {len(ghost_df):,} rows")
log(f"  CLEAN_ITEMS: {len(clean_df):,} rows")


# ── 5. Write cleaned weekly Excel (input for scripts 15/16/17) ───────────────
log("Writing cleaned weekly source file...")

clean_wb = Workbook()
clean_wb.remove(clean_wb.active)

total_removed = 0
for sheet, (df_orig, week_cols) in sheet_data.items():
    df_orig["ITEM_ID"] = pd.to_numeric(df_orig["ITEM_ID"], errors="coerce").fillna(0).astype(int)
    before = len(df_orig)
    df_clean = df_orig[df_orig["ITEM_ID"].isin(clean_ids)].copy()
    removed  = before - len(df_clean)
    total_removed += removed

    # Drop the helper columns added earlier
    drop_cols = [c for c in ["_TOTAL", "_SELL_WKS", "_ACCOUNT", "GRAND_TOTAL"] if c in df_clean.columns]
    df_clean  = df_clean.drop(columns=drop_cols)

    avail_info = [c for c in INFO_COLS if c in df_clean.columns]
    out_cols   = avail_info + week_cols
    df_clean   = df_clean[out_cols].sort_values("ITEM_ID").reset_index(drop=True)

    ws = clean_wb.create_sheet(title=sheet[:31])
    for ci, h in enumerate(out_cols, 1):
        cell = ws.cell(row=1, column=ci, value=str(h))
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for ri, row_vals in enumerate(df_clean.itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

    log(f"  {sheet:<30} {before:>4} → {len(df_clean):>4} items  (removed {removed})")

clean_wb.save(OUT_CLEAN)
log(f"Saved: {OUT_CLEAN}  (total removed: {total_removed} item-account rows)")
log(f"Run scripts 15, 16, 17 using the clean file as input.")
