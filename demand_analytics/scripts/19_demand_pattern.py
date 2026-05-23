"""
19_demand_pattern.py

Classifies every item by demand pattern using the ADI × CV² (Syntetos-Boylan) matrix.

  ADI  = average demand interval = total_weeks / weeks_with_demand_>0
  CV²  = (std of non-zero demand quantities)² / (mean of non-zero demand quantities)²

               CV² ≤ 0.49        CV² > 0.49
  ADI ≤ 1.32   Smooth            Erratic
  ADI >  1.32  Intermittent      Lumpy

Smooth + Erratic items have regular-enough demand to produce meaningful seasonal
signals.  Lumpy + Intermittent items are kept in velocity/lifecycle analyses but
excluded from the PRODFAM seasonality index computation in script 15.

Outputs
-------
output/top20_demand_pattern.xlsx   — audit file (5 sheets: ALL + one per class)
output/demand_pattern_ids.csv      — ITEM_ID, ADI, CV2, DEMAND_PATTERN (merge key)
"""

import os, datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Thresholds ────────────────────────────────────────────────────────────────
ADI_THRESH = 1.32   # Syntetos-Boylan: avg weeks between demand events
CV2_THRESH = 0.49   # Syntetos-Boylan: squared coefficient of variation

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE  = os.path.join(BASE, "output", "top20_accounts_weekly_clean.xlsx")
OUT_XL  = os.path.join(BASE, "output", "top20_demand_pattern.xlsx")
OUT_CSV = os.path.join(BASE, "output", "demand_pattern_ids.csv")

INFO_COLS = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CENTER      = Alignment(horizontal="center")
BOLD_WHITE  = Font(bold=True, color="FFFFFF")

PATTERN_FILLS = {
    "Smooth":       PatternFill("solid", fgColor="00B050"),
    "Erratic":      PatternFill("solid", fgColor="FFBF00"),
    "Intermittent": PatternFill("solid", fgColor="FF7F00"),
    "Lumpy":        PatternFill("solid", fgColor="FF0000"),
}


def log(msg):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}")


def style_ws(ws):
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── 1. Load all sheets, aggregate to item × week totals ───────────────────────
log("Loading clean weekly data...")
xl = pd.ExcelFile(SOURCE)

frames = []
info_rows = {}

for sheet in xl.sheet_names:
    df = xl.parse(sheet)
    week_cols = [c for c in df.columns
                 if isinstance(c, str) and c[0].isdigit() and "W" in c]

    for _, row in df.iterrows():
        iid = int(pd.to_numeric(row.get("ITEM_ID", 0), errors="coerce") or 0)
        if iid and iid not in info_rows:
            info_rows[iid] = {c: row.get(c, "") for c in INFO_COLS if c != "ITEM_ID"}

    keep = [c for c in INFO_COLS if c in df.columns] + week_cols
    frames.append(df[keep])

combined = pd.concat(frames, ignore_index=True)
combined["ITEM_ID"] = pd.to_numeric(combined["ITEM_ID"], errors="coerce").fillna(0).astype(int)

all_week_cols = [c for c in combined.columns
                 if isinstance(c, str) and c[0].isdigit() and "W" in c]

log(f"Accounts: {len(xl.sheet_names)}  |  Items: {combined['ITEM_ID'].nunique():,}  |  Week cols: {len(all_week_cols)}")

# ── 2. Sum across accounts per item ──────────────────────────────────────────
log("Aggregating weekly demand per item...")
item_weekly = combined.groupby("ITEM_ID")[all_week_cols].sum().reset_index()
demand_mat  = item_weekly[all_week_cols].to_numpy(dtype=float)
total_periods = demand_mat.shape[1]

# ── 3. Compute ADI and CV² ────────────────────────────────────────────────────
log("Computing ADI and CV²...")

nz_mask  = demand_mat > 0
nz_weeks = nz_mask.sum(axis=1)

adi = np.where(nz_weeks > 0, total_periods / nz_weeks, np.inf)

cv2 = np.zeros(len(item_weekly))
for i in range(len(item_weekly)):
    nz = demand_mat[i][nz_mask[i]]
    if len(nz) >= 2:
        cv2[i] = np.var(nz, ddof=1) / (np.mean(nz) ** 2)

# ── 4. Classify ───────────────────────────────────────────────────────────────
def classify(a, c):
    if a <= ADI_THRESH and c <= CV2_THRESH: return "Smooth"
    if a <= ADI_THRESH and c >  CV2_THRESH: return "Erratic"
    if a >  ADI_THRESH and c <= CV2_THRESH: return "Intermittent"
    return "Lumpy"

item_weekly["NON_ZERO_WEEKS"] = nz_weeks.astype(int)
item_weekly["TOTAL_WEEKS"]    = total_periods
item_weekly["ADI"]            = np.round(adi, 2)
item_weekly["CV2"]            = np.round(cv2, 3)
item_weekly["DEMAND_PATTERN"] = [classify(a, c) for a, c in zip(adi, cv2)]

# ── 5. Merge item info ────────────────────────────────────────────────────────
info_df = pd.DataFrame([{"ITEM_ID": iid, **v} for iid, v in info_rows.items()])
info_df["ITEM_ID"] = info_df["ITEM_ID"].astype(int)

result = item_weekly[["ITEM_ID", "NON_ZERO_WEEKS", "TOTAL_WEEKS",
                       "ADI", "CV2", "DEMAND_PATTERN"]].merge(info_df, on="ITEM_ID", how="left")

col_order = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM",
             "NON_ZERO_WEEKS", "TOTAL_WEEKS", "ADI", "CV2", "DEMAND_PATTERN"]
result = result[col_order].sort_values(["DEMAND_PATTERN", "ADI"],
                                        ascending=[True, False]).reset_index(drop=True)

counts = result["DEMAND_PATTERN"].value_counts()
log(f"Total items : {len(result):,}")
for pat in ["Smooth", "Erratic", "Intermittent", "Lumpy"]:
    log(f"  {pat:<13}: {counts.get(pat, 0):>4}")

# ── 6. Save CSV ───────────────────────────────────────────────────────────────
result[["ITEM_ID", "ADI", "CV2", "DEMAND_PATTERN"]].to_csv(OUT_CSV, index=False)
log(f"Saved: {OUT_CSV}")

# ── 7. Save Excel ─────────────────────────────────────────────────────────────
log("Writing Excel report...")
pat_col_idx = col_order.index("DEMAND_PATTERN") + 1

wb = Workbook()
wb.remove(wb.active)

sheets = [
    ("ALL_ITEMS",    result),
    ("SMOOTH",       result[result["DEMAND_PATTERN"] == "Smooth"].reset_index(drop=True)),
    ("ERRATIC",      result[result["DEMAND_PATTERN"] == "Erratic"].reset_index(drop=True)),
    ("INTERMITTENT", result[result["DEMAND_PATTERN"] == "Intermittent"].reset_index(drop=True)),
    ("LUMPY",        result[result["DEMAND_PATTERN"] == "Lumpy"].reset_index(drop=True)),
]

for title, data in sheets:
    ws = wb.create_sheet(title=title)
    for ci, h in enumerate(col_order, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    for ri, row_vals in enumerate(data[col_order].itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == pat_col_idx:
                cell.fill      = PATTERN_FILLS.get(str(val), PatternFill())
                cell.font      = BOLD_WHITE
                cell.alignment = CENTER
    style_ws(ws)

wb.save(OUT_XL)
log(f"Saved: {OUT_XL}")
log("Run order: 19 → 15 → 16 → 17 → build_velocity_parquet.py")
