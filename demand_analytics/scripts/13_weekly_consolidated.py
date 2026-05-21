"""
13_weekly_consolidated.py

Converts top20_accounts_consolidated.xlsx (daily date columns)
into a weekly version by summing all daily quantities within each ISO week.

Output structure (mirrors the source file):
  Rows    = items  (ITEM_ID, ITEM_DESCR, BRAND, VENDOR, PRODFAM)
  Columns = weeks  (2023-W01, 2023-W02, … 2026-W13) + GRAND_TOTAL
  Sheets  = one per account (20 sheets)

Output: account_consolidator/output/top20_accounts_weekly.xlsx

Run from Dataprocessing/:
    python3 13_weekly_consolidated.py
"""

import os
import time
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE, "output", "top20_accounts_consolidated.xlsx")
OUT_XLSX   = os.path.join(BASE, "output", "top20_accounts_weekly.xlsx")

INFO_COLS  = ["ITEM_ID", "ITEM_DESCR", "BRAND", "VENDOR", "PRODFAM"]

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
ZERO_FILL   = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FONT  = Font(bold=True, size=10)
CENTER      = Alignment(horizontal="center")


def log(msg: str, level: str = "INFO"):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [{level}] {msg}")


def section(title: str):
    log("─" * 60)
    log(f"  STEP: {title}")
    log("─" * 60)


section("Load Source & Build Weekly Columns")
xl = pd.ExcelFile(SOURCE)
log(f"Source : {SOURCE}")
log(f"Sheets : {xl.sheet_names}")

# Pre-build date → WEEK_LABEL mapping from the first sheet's date columns
_sample = xl.parse(xl.sheet_names[0], nrows=0)
all_date_cols = [c for c in _sample.columns if c not in INFO_COLS and c != "GRAND_TOTAL"]
date_index    = pd.to_datetime(all_date_cols, errors="coerce")
week_labels   = date_index.strftime("%Y-W%W")          # e.g. "2023-W01"
date_to_week  = dict(zip(all_date_cols, week_labels))  # "2023-01-04" → "2023-W01"

all_weeks = sorted(set(week_labels))
log(f"Daily date columns : {len(all_date_cols)}")
log(f"Unique week buckets: {len(all_weeks)}  ({all_weeks[0]} → {all_weeks[-1]})")


section("Process Each Sheet & Write Excel")
t0_total = time.time()

wb = openpyxl.Workbook()
wb.remove(wb.active)

for sheet in xl.sheet_names:
    t0 = time.time()
    df_wide = xl.parse(sheet)

    avail_info = [c for c in INFO_COLS if c in df_wide.columns]
    date_cols  = [c for c in df_wide.columns if c in date_to_week]

    # Melt daily → long, map to week label, sum per item per week
    df_long = df_wide[avail_info + date_cols].melt(
        id_vars=avail_info, var_name="DATE", value_name="QTY"
    )
    df_long["QTY"]        = pd.to_numeric(df_long["QTY"], errors="coerce").fillna(0).astype(int)
    df_long["WEEK_LABEL"] = df_long["DATE"].map(date_to_week)

    weekly = (
        df_long
        .groupby(avail_info + ["WEEK_LABEL"], as_index=False)["QTY"]
        .sum()
    )

    # Pivot back to wide: rows = items, columns = weeks
    pivot = weekly.pivot_table(
        index=avail_info, columns="WEEK_LABEL", values="QTY",
        aggfunc="sum", fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    # Ensure all weeks present even if this account had no data for some
    for w in all_weeks:
        if w not in pivot.columns:
            pivot[w] = 0

    week_cols = sorted([c for c in pivot.columns if c not in avail_info])
    pivot     = pivot[avail_info + week_cols]

    pivot["GRAND_TOTAL"] = pivot[week_cols].sum(axis=1)
    pivot = pivot.sort_values("GRAND_TOTAL", ascending=False).reset_index(drop=True)

    all_cols       = avail_info + week_cols + ["GRAND_TOTAL"]
    grand_col_idx  = len(all_cols)          # 1-based index of GRAND_TOTAL column
    info_col_count = len(avail_info)

    # ── Write sheet ───────────────────────────────────────────────────────────
    ws = wb.create_sheet(title=sheet[:31])

    # Header row
    for ci, h in enumerate(all_cols, 1):
        cell           = ws.cell(row=1, column=ci, value=str(h))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # Data rows
    for ri, row in enumerate(pivot[all_cols].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == grand_col_idx:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            elif ci > info_col_count and (val == 0 or val == "0"):
                cell.fill = ZERO_FILL

    # Freeze panes after info columns
    ws.freeze_panes = ws.cell(row=2, column=info_col_count + 1)

    # Column widths
    for ci, h in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 2, 30)

    log(f"  {sheet:<30} {len(pivot):>5} items × {len(week_cols):>3} weeks  ({time.time()-t0:.1f}s)")

section("Save Workbook")
wb.save(OUT_XLSX)
size_kb = os.path.getsize(OUT_XLSX) / 1024
log(f"Saved : {OUT_XLSX}")
log(f"Size  : {size_kb:,.1f} KB")
log(f"Total : {time.time()-t0_total:.1f}s")
